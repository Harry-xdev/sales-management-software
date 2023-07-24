import json
import openpyxl
from openpyxl import load_workbook
import os
import time

def create_txt(filename):
	with open(filename, "w") as file:
		pass
	print(f"{filename} created.")


def create_xlsx(filename):
	workbook = openpyxl.Workbook()
	worksheet = workbook.active
	worksheet.title = "Sheet1"
	workbook.save(filename)
	print(f"{filename} created.")


def load_txt(filename):
	with open(filename, "r") as file:
		data_json = file.read().strip()
		data_dict = json.loads(data_json)
	return data_dict


def export_txt(filename, data):
	data_json = json.dumps(data)
	with open(filename, "w") as file:
		file.write(data_json)
		print(f"{filename} saved.")
		
def export_item_txt(filename, data):
	
	old_data = load_txt(filename)
	data_dict = [i.__dict__ for i in data]
	combine = data_dict + old_data
	combine_json = json.dumps(combine)

	with open(filename, "w") as file:
		file.write(combine_json)
		print(f"{filename} saved.")

def export_item_xlsx(name_xlsx, data_txt):
	
	# set column txt
	data = load_txt(data_txt)
	for i in range(len(data)):
		data[i]["A"] = data[i].pop("id")
		data[i]["B"] = data[i].pop("name")
		data[i]["C"] = data[i].pop("price")
	print(data)
	
	# save data to xlsx
	if not os.path.exists(name_xlsx):
		print(f"creating {name_xlsx} file...")
		time.sleep(1)
		create_xlsx(name_xlsx)
	
	# set column xlsx
	workbook = load_workbook(filename=name_xlsx)
	worksheet = workbook["Sheet1"]
	worksheet["A1"] = "id"
	worksheet["B1"] = "name"
	worksheet["C1"] = "price"
	
	for i in data:
		worksheet.append(i)
	workbook.save(name_xlsx)
	print(f"{name_xlsx} saved.")


def export_table_xlsx(name_xlsx, data_txt):
	data = load_txt(data_txt)
	# set column txt
	for i in range(len(data)):
		data[i]["A"] = data[i].pop("id")
		data[i]["B"] = data[i].pop("name")
		data[i]["C"] = data[i].pop("cost")
		data[i]["D"] = data[i].pop("start")
		data[i]["E"] = data[i].pop("end")
		data[i]["F"] = data[i].pop("a")
		data[i]["G"] = data[i].pop("b")
		data[i]["H"] = data[i].pop("hour")
		data[i]["I"] = data[i].pop("item")
		data[i]["J"] = data[i].pop("total_cost")
		data[i]["K"] = data[i].pop("is_active")
	
	# save data to xlsx
	if not os.path.exists(name_xlsx):
		time.sleep(1)
		create_xlsx(name_xlsx)
		print(f"creating {name_xlsx} file...")
	
	# set column xlsx
	workbook = load_workbook(filename=name_xlsx)
	worksheet = workbook["Sheet1"]
	worksheet["A1"] = "id"
	worksheet["B1"] = "name"
	worksheet["C1"] = "cost"
	worksheet["D1"] = "start"
	worksheet["E1"] = "end"
	worksheet["F1"] = "a"
	worksheet["G1"] = "b"
	worksheet["H1"] = "hour"
	worksheet["I1"] = "item"
	worksheet["J1"] = "total_cost"
	worksheet["K1"] = "is_active"
	
	# append data from txt to xlsx
	for i in data:
		worksheet.append(i)
	workbook.save(name_xlsx)
	print(f"{name_xlsx} saved.")

	



